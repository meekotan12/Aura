"""Use: Handles bulk student import API endpoints.
Where to use: Use this through the FastAPI app when the frontend or an API client needs bulk student import features.
Role: Router layer. It receives HTTP requests, checks access rules, and returns API responses.
"""

from __future__ import annotations

import io
import json
import logging
import os
import uuid
from pathlib import Path
from zipfile import BadZipFile

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile, status
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.security import get_current_user_with_roles, has_any_role
from app.core.dependencies import get_db
from app.models.associations import program_department_association
from app.models.department import Department
from app.models.import_job import BulkImportJob
from app.models.program import Program
from app.models.school import SchoolAuditLog
from app.models.user import User
from app.repositories.import_repository import ImportRepository
from app.schemas.import_job import (
    ImportErrorItem,
    ImportJobCreateResponse,
    ImportJobStatusResponse,
    ImportPreviewResponse,
    ImportPreviewRow,
    RetryFailedRowsRequest,
)
from app.services.import_file_service import (
    is_supported_import_file,
    load_tabular_rows_from_bytes,
    normalize_upload_to_csv_bytes,
)
from app.services.import_validation_service import (
    EXPECTED_HEADERS,
    HeaderValidationError,
    ValidationContext,
    suggest_fixes,
    validate_and_transform_row,
    validate_headers,
)
from app.services.student_import_service import StudentImportService
from app.workers.celery_app import celery_app

router = APIRouter(prefix="/api/admin", tags=["admin-import"])
logger = logging.getLogger(__name__)


def get_current_admin_or_school_it(
    current_user: User = Depends(get_current_user_with_roles),
) -> User:
    """Allow only Admin or Campus Admin users to use import routes."""
    if not has_any_role(current_user, ["admin", "campus_admin"]):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin or Campus Admin privileges required",
        )
    return current_user


def _append_import_audit_log(
    db: Session,
    *,
    current_user: User,
    status_value: str,
    details: dict,
    action: str = "student_bulk_import_attempt",
) -> None:
    """Save an audit log entry for this import action when the user has a school."""
    school_id = getattr(current_user, "school_id", None)
    if school_id is None:
        return

    db.add(
        SchoolAuditLog(
            school_id=school_id,
            actor_user_id=current_user.id,
            action=action,
            status=status_value,
            details=json.dumps(details, default=str),
        )
    )


def _ensure_user_school(current_user: User) -> int:
    """Every import job belongs to one school, so stop if the user has no school."""
    school_id = getattr(current_user, "school_id", None)
    if school_id is None:
        raise HTTPException(status_code=403, detail="User is not assigned to a school")
    return school_id


def _validate_upload_basics(
    *,
    file: UploadFile,
    current_user: User,
    db: Session,
    settings,
) -> tuple[str, int]:
    """Check the file name, type, and size before we read the import file."""
    filename = (file.filename or "").strip()
    if not filename:
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="failed",
            details={"reason": "missing file name"},
        )
        db.commit()
        raise HTTPException(status_code=400, detail="File name is required")
    if not is_supported_import_file(filename):
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="failed",
            details={"reason": "unsupported file extension", "filename": filename},
        )
        db.commit()
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx files are allowed")

    file.file.seek(0, os.SEEK_END)
    size_bytes = file.file.tell()
    file.file.seek(0)
    if size_bytes <= 0:
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="failed",
            details={"reason": "empty upload", "filename": filename},
        )
        db.commit()
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    max_size_bytes = settings.import_max_file_size_mb * 1024 * 1024
    if size_bytes > max_size_bytes:
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="failed",
            details={
                "reason": "file size exceeded",
                "filename": filename,
                "size_bytes": size_bytes,
                "max_size_bytes": max_size_bytes,
            },
        )
        db.commit()
        raise HTTPException(
            status_code=413,
            detail=f"File size exceeds limit of {settings.import_max_file_size_mb} MB",
        )
    return filename, size_bytes


def _queue_import_job_from_file_bytes(
    *,
    db: Session,
    background_tasks: BackgroundTasks,
    settings,
    current_user: User,
    filename: str,
    file_bytes: bytes,
    size_bytes: int,
    retried_from_job_id: str | None = None,
) -> ImportJobCreateResponse:
    """Normalize an uploaded import file first, then queue the background import job."""
    storage_dir = Path(settings.import_storage_dir) / "uploads"
    storage_dir.mkdir(parents=True, exist_ok=True)
    _, normalized_file_bytes = normalize_upload_to_csv_bytes(
        filename=filename,
        file_bytes=file_bytes,
    )
    stored_file_path = storage_dir / f"{uuid.uuid4()}.csv"
    stored_file_path.write_bytes(normalized_file_bytes)
    return _queue_import_job_from_stored_path(
        db=db,
        background_tasks=background_tasks,
        settings=settings,
        current_user=current_user,
        filename=filename,
        stored_file_path=str(stored_file_path),
        size_bytes=size_bytes if size_bytes > 0 else len(normalized_file_bytes),
        retried_from_job_id=retried_from_job_id,
    )


def _queue_import_job_from_stored_path(
    *,
    db: Session,
    background_tasks: BackgroundTasks,
    settings,
    current_user: User,
    filename: str,
    stored_file_path: str,
    size_bytes: int,
    retried_from_job_id: str | None = None,
    preview_token: str | None = None,
) -> ImportJobCreateResponse:
    """Create the job row around a server-side stored import payload and queue the worker."""
    repo = ImportRepository(db)
    recent_job_count = repo.count_recent_jobs(
        created_by_user_id=current_user.id,
        window_seconds=settings.import_rate_limit_window_seconds,
    )
    if recent_job_count >= settings.import_rate_limit_count:
        _append_import_audit_log(
            db,
            current_user=current_user,
            status_value="rate_limited",
            details={
                "reason": "rate limit exceeded",
                "filename": filename,
                "window_seconds": settings.import_rate_limit_window_seconds,
            },
        )
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_429_TOO_MANY_REQUESTS,
            detail="Too many import requests. Please wait before uploading again.",
        )

    job_id = str(uuid.uuid4())
    target_school_id = _ensure_user_school(current_user)

    job = BulkImportJob(
        id=job_id,
        created_by_user_id=current_user.id,
        target_school_id=target_school_id,
        status="pending",
        original_filename=filename,
        stored_file_path=stored_file_path,
    )

    repo.create_job(job)
    audit_details = {
        "job_id": job_id,
        "filename": filename,
        "size_bytes": size_bytes,
        "retried_from_job_id": retried_from_job_id,
    }
    if preview_token:
        audit_details["preview_token"] = preview_token
    _append_import_audit_log(
        db,
        current_user=current_user,
        status_value="queued",
        details=audit_details,
        action="student_bulk_import_retry" if retried_from_job_id else "student_bulk_import_attempt",
    )
    db.commit()
    # Queue the background job after the database row is safely saved.
    try:
        celery_app.send_task("app.workers.tasks.process_student_import_job", args=[job_id])
    except Exception:
        logger.warning(
            "Falling back to in-process import execution for job %s because Celery dispatch failed.",
            job_id,
            exc_info=True,
        )
        background_tasks.add_task(StudentImportService().process_job, job_id)

    return ImportJobCreateResponse(
        job_id=job_id,
        status="pending",
        retried_from_job_id=retried_from_job_id,
    )


def _build_validation_context(db: Session, target_school_id: int) -> ValidationContext:
    """Build simple name-to-id maps used when we validate each row."""
    department_lookup = {
        name.strip().lower(): department_id
        for department_id, name in (
            db.query(Department.id, Department.name)
            .filter(Department.school_id == target_school_id)
            .all()
        )
    }
    course_lookup = {
        name.strip().lower(): program_id
        for program_id, name in (
            db.query(Program.id, Program.name)
            .filter(Program.school_id == target_school_id)
            .all()
        )
    }
    department_course_pairs = {
        (int(department_id), int(program_id))
        for program_id, department_id in db.execute(
            select(
                program_department_association.c.program_id,
                program_department_association.c.department_id,
            )
            .select_from(program_department_association)
            .join(Program, Program.id == program_department_association.c.program_id)
            .join(Department, Department.id == program_department_association.c.department_id)
            .where(
                Program.school_id == target_school_id,
                Department.school_id == target_school_id,
            )
        ).all()
    }
    return ValidationContext(
        target_school_id=target_school_id,
        department_lookup=department_lookup,
        course_lookup=course_lookup,
        department_course_pairs=department_course_pairs,
    )


def _preview_manifest_dir(settings) -> Path:
    return Path(settings.import_storage_dir) / "previews"


def _preview_manifest_path(settings, preview_token: str) -> Path:
    return _preview_manifest_dir(settings) / f"{preview_token}.json"


def _write_preview_manifest(
    *,
    settings,
    preview_token: str,
    manifest: dict,
) -> None:
    preview_dir = _preview_manifest_dir(settings)
    preview_dir.mkdir(parents=True, exist_ok=True)
    _preview_manifest_path(settings, preview_token).write_text(
        json.dumps(manifest, default=str),
        encoding="utf-8",
    )


def _store_preview_manifest(
    *,
    settings,
    current_user: User,
    target_school_id: int,
    filename: str,
    total_rows: int,
    valid_rows: int,
    invalid_rows: int,
    can_commit: bool,
    rows: list[dict],
    error_rows: list[dict],
) -> str:
    preview_token = str(uuid.uuid4())
    manifest = {
        "preview_token": preview_token,
        "created_by_user_id": current_user.id,
        "target_school_id": target_school_id,
        "original_filename": filename,
        "total_rows": total_rows,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "can_commit": can_commit,
        "rows": rows,
        "error_rows": error_rows,
    }
    _write_preview_manifest(
        settings=settings,
        preview_token=preview_token,
        manifest=manifest,
    )
    return preview_token


def _load_preview_manifest(
    *,
    settings,
    preview_token: str,
    current_user: User,
) -> tuple[dict, Path]:
    manifest_path = _preview_manifest_path(settings, preview_token)
    if not manifest_path.exists():
        raise HTTPException(
            status_code=404,
            detail="Approved preview not found. Preview the file again before importing.",
        )

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise HTTPException(
            status_code=400,
            detail="Approved preview data is invalid. Preview the file again before importing.",
        ) from exc

    if manifest.get("created_by_user_id") != current_user.id:
        raise HTTPException(status_code=404, detail="Approved preview not found")

    target_school_id = _ensure_user_school(current_user)
    if manifest.get("target_school_id") != target_school_id:
        raise HTTPException(status_code=404, detail="Approved preview not found")

    if not isinstance(manifest.get("rows"), list):
        raise HTTPException(
            status_code=400,
            detail="Approved preview data is incomplete. Preview the file again before importing.",
        )

    return manifest, manifest_path


def _build_preview_rows_from_manifest_rows(manifest_rows: list[dict]) -> list[ImportPreviewRow]:
    preview_rows: list[ImportPreviewRow] = []
    sorted_rows = sorted(
        (row for row in manifest_rows if isinstance(row, dict)),
        key=lambda item: int(item.get("row_number") or 0),
    )
    for row in sorted_rows[:200]:
        row_number = int(row.get("row_number") or 0)
        raw_row_data = row.get("raw_row_data") if isinstance(row.get("raw_row_data"), dict) else None
        preview_rows.append(
            ImportPreviewRow(
                row=row_number,
                status="valid",
                errors=[],
                suggestions=[],
                row_data=raw_row_data,
            )
        )
    return preview_rows


def _build_preview_response_from_manifest(manifest: dict) -> ImportPreviewResponse:
    manifest_rows = manifest.get("rows") if isinstance(manifest.get("rows"), list) else []
    preview_token = manifest.get("preview_token")
    return ImportPreviewResponse(
        filename=str(manifest.get("original_filename") or "student_import.xlsx"),
        total_rows=int(manifest.get("total_rows") or len(manifest_rows)),
        valid_rows=int(manifest.get("valid_rows") or 0),
        invalid_rows=int(manifest.get("invalid_rows") or 0),
        can_commit=bool(manifest.get("can_commit")),
        preview_token=str(preview_token) if preview_token else None,
        rows=_build_preview_rows_from_manifest_rows(manifest_rows),
    )


def _find_persistent_row_conflicts(db: Session, rows: list[dict]) -> dict[int, list[str]]:
    if not rows:
        return {}

    repo = ImportRepository(db)
    existing_emails = repo.existing_emails(row["email"] for row in rows)
    existing_pairs = repo.existing_school_student_pairs(
        (row["school_id"], row["student_id"]) for row in rows
    )

    conflicts: dict[int, list[str]] = {}
    for row in rows:
        row_errors: list[str] = []
        if row["email"] in existing_emails:
            row_errors.append("Email already exists")
        if (row["school_id"], row["student_id"]) in existing_pairs:
            row_errors.append("Duplicate Student_ID within School_ID")
        if row_errors:
            conflicts[int(row["row_number"])] = row_errors

    return conflicts


def _build_retry_workbook_bytes(*, sheet_title: str, row_payloads: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    sheet.append(EXPECTED_HEADERS)
    for row_data in row_payloads:
        sheet.append([str(row_data.get(header, "")) for header in EXPECTED_HEADERS])

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.read()


def _build_preview_error_report_bytes(error_payloads: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Preview Errors"
    sheet.append(EXPECTED_HEADERS + ["Error"])
    for item in error_payloads:
        row_data = item.get("row_data") if isinstance(item.get("row_data"), dict) else {}
        error_message = "; ".join(item.get("errors") or []) or "Unknown preview error"
        sheet.append([str(row_data.get(header, "")) for header in EXPECTED_HEADERS] + [error_message])

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    output.seek(0)
    return output.read()


@router.get("/import-students/template")
def download_import_students_template(
    current_user: User = Depends(get_current_admin_or_school_it),
):
    """Download the Excel template that users should fill in before importing."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    sheet.append(EXPECTED_HEADERS)
    # Add one sample row so users can see the expected format right away.
    sheet.append(
        [
            "STU-00001",
            "student1@example.edu",
            "Doe",
            "Jane",
            "A",
            "Computer Science",
            "BS Computer Science",
        ]
    )

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="student_import_template.xlsx"'},
    )


@router.post("/import-students/preview", response_model=ImportPreviewResponse)
def preview_import_students(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    """Read the uploaded import file and show row errors before starting the real import."""
    settings = get_settings()
    filename, _ = _validate_upload_basics(
        file=file,
        current_user=current_user,
        db=db,
        settings=settings,
    )
    target_school_id = _ensure_user_school(current_user)
    file.file.seek(0)
    file_bytes = file.file.read()

    try:
        tabular_rows = load_tabular_rows_from_bytes(
            filename=filename,
            file_bytes=file_bytes,
        )
        total_rows = max(len(tabular_rows) - 1, 0)
        context = _build_validation_context(db, target_school_id)

        # The first row must match the template headers exactly.
        header_row = tabular_rows[0] if tabular_rows else None
        if header_row is None:
            return ImportPreviewResponse(
                filename=filename,
                total_rows=0,
                valid_rows=0,
                invalid_rows=0,
                can_commit=False,
                rows=[
                    ImportPreviewRow(
                        row=1,
                        status="failed",
                        errors=["Missing header row"],
                        suggestions=["Download and use the latest template before importing."],
                        row_data=None,
                    )
                ],
            )

        try:
            validate_headers(header_row)
        except HeaderValidationError as exc:
            message = str(exc)
            return ImportPreviewResponse(
                filename=filename,
                total_rows=total_rows,
                valid_rows=0,
                invalid_rows=max(total_rows, 1),
                can_commit=False,
                rows=[
                    ImportPreviewRow(
                        row=1,
                        status="failed",
                        errors=[message],
                        suggestions=suggest_fixes([message]),
                        row_data=None,
                    )
                ],
            )

        valid_rows = 0
        invalid_rows = 0
        approved_rows: list[dict] = []
        error_rows: list[dict] = []
        preview_rows: list[ImportPreviewRow] = []
        preview_rows_by_number: dict[int, ImportPreviewRow] = {}

        for row_number, row_values in enumerate(tabular_rows[1:], start=2):
            transformed, row_errors, row_data = validate_and_transform_row(
                row_number=row_number,
                row_values=row_values,
                context=context,
            )
            if row_errors:
                invalid_rows += 1
                status_value = "failed"
                errors = row_errors
                suggestions = suggest_fixes(row_errors)
                error_rows.append(
                    {
                        "row": row_number,
                        "row_data": row_data,
                        "errors": row_errors,
                    }
                )
            else:
                valid_rows += 1
                status_value = "valid"
                errors = []
                suggestions = []
                transformed["raw_row_data"] = row_data
                approved_rows.append(transformed)

            # Keep the preview response small even if the upload is very large.
            if len(preview_rows) < 200:
                preview_row = ImportPreviewRow(
                    row=row_number,
                    status=status_value,
                    errors=errors,
                    suggestions=suggestions,
                    row_data=row_data if transformed is None else row_data,
                )
                preview_rows.append(preview_row)
                preview_rows_by_number[row_number] = preview_row

        persistent_conflicts = _find_persistent_row_conflicts(db, approved_rows)
        if persistent_conflicts:
            valid_rows -= len(persistent_conflicts)
            invalid_rows += len(persistent_conflicts)

            filtered_rows: list[dict] = []
            for row in approved_rows:
                row_errors = persistent_conflicts.get(int(row["row_number"]))
                if row_errors:
                    preview_row = preview_rows_by_number.get(int(row["row_number"]))
                    if preview_row is not None:
                        preview_row.status = "failed"
                        preview_row.errors = row_errors
                        preview_row.suggestions = suggest_fixes(row_errors)
                    error_rows.append(
                        {
                            "row": int(row["row_number"]),
                            "row_data": row.get("raw_row_data"),
                            "errors": row_errors,
                        }
                    )
                    continue
                filtered_rows.append(row)
            approved_rows = filtered_rows

        can_commit = total_rows > 0 and invalid_rows == 0
        preview_token = _store_preview_manifest(
            settings=settings,
            current_user=current_user,
            target_school_id=target_school_id,
            filename=filename,
            total_rows=total_rows,
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            can_commit=can_commit,
            rows=approved_rows,
            error_rows=error_rows,
        )

        return ImportPreviewResponse(
            filename=filename,
            total_rows=total_rows,
            valid_rows=valid_rows,
            invalid_rows=invalid_rows,
            can_commit=can_commit,
            preview_token=preview_token,
            rows=preview_rows,
        )
    except (InvalidFileException, BadZipFile, OSError, UnicodeDecodeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=f"Invalid import file: {exc}") from exc


@router.post("/import-students", response_model=ImportJobCreateResponse)
def import_students(
    background_tasks: BackgroundTasks,
    preview_token: str | None = Form(default=None),
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    """Queue a preview-approved import job from a stored preview manifest."""
    settings = get_settings()
    if preview_token:
        manifest, manifest_path = _load_preview_manifest(
            settings=settings,
            preview_token=preview_token,
            current_user=current_user,
        )
        if not bool(manifest.get("can_commit")):
            raise HTTPException(
                status_code=400,
                detail="Preview still has invalid rows. Fix them before importing.",
            )
        filename = str(manifest.get("original_filename") or "student_import.xlsx")
        size_bytes = manifest_path.stat().st_size if manifest_path.exists() else 0
        return _queue_import_job_from_stored_path(
            db=db,
            background_tasks=background_tasks,
            settings=settings,
            current_user=current_user,
            filename=filename,
            stored_file_path=str(manifest_path),
            size_bytes=size_bytes,
            preview_token=preview_token,
        )

    raise HTTPException(status_code=400, detail="Preview the file first before importing.")


@router.get("/import-preview-errors/{preview_token}/download")
def download_preview_errors(
    preview_token: str,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    """Download an Excel report of rows that failed during preview validation."""
    settings = get_settings()
    manifest, _ = _load_preview_manifest(
        settings=settings,
        preview_token=preview_token,
        current_user=current_user,
    )
    error_rows = manifest.get("error_rows")
    if not isinstance(error_rows, list) or not error_rows:
        raise HTTPException(status_code=404, detail="No preview errors available to download")

    original_filename = str(manifest.get("original_filename") or "student_import.xlsx")
    report_bytes = _build_preview_error_report_bytes(error_rows)
    download_name = f"preview_errors_{Path(original_filename).stem}.xlsx"

    return StreamingResponse(
        io.BytesIO(report_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


@router.get("/import-preview-errors/{preview_token}/retry-download")
def download_preview_retry_file(
    preview_token: str,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    """Download an Excel file containing only preview-failed rows for correction and re-upload."""
    settings = get_settings()
    manifest, _ = _load_preview_manifest(
        settings=settings,
        preview_token=preview_token,
        current_user=current_user,
    )
    error_rows = manifest.get("error_rows")
    if not isinstance(error_rows, list) or not error_rows:
        raise HTTPException(status_code=404, detail="No preview errors available to retry")

    retry_row_payloads = [
        item["row_data"]
        for item in error_rows
        if isinstance(item, dict) and isinstance(item.get("row_data"), dict)
    ]
    if not retry_row_payloads:
        raise HTTPException(
            status_code=404,
            detail="No retryable row payloads found for this preview",
        )

    original_filename = str(manifest.get("original_filename") or "student_import.xlsx")
    retry_bytes = _build_retry_workbook_bytes(
        sheet_title="Students-Retry",
        row_payloads=retry_row_payloads,
    )
    download_name = f"preview_retry_{Path(original_filename).name}"

    return StreamingResponse(
        io.BytesIO(retry_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


@router.post(
    "/import-preview-errors/{preview_token}/remove-invalid",
    response_model=ImportPreviewResponse,
)
def remove_invalid_preview_rows(
    preview_token: str,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    """Keep only preview-approved rows so the preview can proceed to import."""
    settings = get_settings()
    manifest, _ = _load_preview_manifest(
        settings=settings,
        preview_token=preview_token,
        current_user=current_user,
    )
    approved_rows = manifest.get("rows")
    if not isinstance(approved_rows, list):
        raise HTTPException(
            status_code=400,
            detail="Preview data is incomplete. Preview the file again before importing.",
        )
    if not approved_rows:
        raise HTTPException(
            status_code=400,
            detail="No valid preview rows available to keep.",
        )

    cleaned_manifest = dict(manifest)
    cleaned_manifest["total_rows"] = len(approved_rows)
    cleaned_manifest["valid_rows"] = len(approved_rows)
    cleaned_manifest["invalid_rows"] = 0
    cleaned_manifest["can_commit"] = True
    cleaned_manifest["error_rows"] = []
    _write_preview_manifest(
        settings=settings,
        preview_token=preview_token,
        manifest=cleaned_manifest,
    )

    _append_import_audit_log(
        db,
        current_user=current_user,
        status_value="preview_cleaned",
        details={
            "preview_token": preview_token,
            "kept_rows": len(approved_rows),
            "dropped_rows": int(manifest.get("invalid_rows") or 0),
            "filename": str(manifest.get("original_filename") or "student_import.xlsx"),
        },
        action="student_bulk_import_preview_cleaned",
    )
    db.commit()

    return _build_preview_response_from_manifest(cleaned_manifest)


@router.post("/import-students/retry-failed/{job_id}", response_model=ImportJobCreateResponse)
def retry_failed_rows(
    job_id: str,
    background_tasks: BackgroundTasks,
    payload: RetryFailedRowsRequest | None = None,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    """Build a new Excel file from failed rows and queue a retry job."""
    repo = ImportRepository(db)
    parent_job = repo.get_job(job_id)
    if not parent_job or parent_job.created_by_user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Import job not found")

    errors = repo.fetch_errors(job_id, limit=10000)
    if not errors:
        raise HTTPException(status_code=400, detail="No failed rows available to retry")

    selected_rows = set((payload.row_numbers if payload else []) or [])
    retry_row_payloads: list[dict] = []
    for item in errors:
        if selected_rows and item.row_number not in selected_rows:
            continue
        if isinstance(item.row_data, dict):
            retry_row_payloads.append(item.row_data)

    if not retry_row_payloads:
        raise HTTPException(
            status_code=400,
            detail="No retryable row payloads found for the selected rows",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students-Retry"
    sheet.append(EXPECTED_HEADERS)
    # Rebuild the failed rows in the same column order as the template.
    for row_data in retry_row_payloads:
        sheet.append([str(row_data.get(header, "")) for header in EXPECTED_HEADERS])

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    file_bytes = output.read()

    retry_filename = f"retry_{parent_job.original_filename or 'student_import.xlsx'}"
    settings = get_settings()
    return _queue_import_job_from_file_bytes(
        db=db,
        background_tasks=background_tasks,
        settings=settings,
        current_user=current_user,
        filename=retry_filename,
        file_bytes=file_bytes,
        size_bytes=len(file_bytes),
        retried_from_job_id=job_id,
    )


@router.get("/import-status/{job_id}", response_model=ImportJobStatusResponse)
def get_import_status(
    job_id: str,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    """Return job progress, row errors, and the failed-row report link when ready."""
    repo = ImportRepository(db)
    job = repo.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    if job.created_by_user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Import job not found")

    percentage = 0.0
    # Avoid division by zero when the worker has not counted rows yet.
    if job.total_rows > 0:
        percentage = round((job.processed_rows / job.total_rows) * 100, 2)

    errors = [
        ImportErrorItem(row=item.row_number, error=item.error_message)
        for item in repo.fetch_errors(job_id, limit=5000)
    ]

    failed_report_download_url = None
    if job.failed_report_path:
        failed_report_download_url = f"/api/admin/import-errors/{job_id}/download"

    return ImportJobStatusResponse(
        job_id=job.id,
        state=job.status,
        total_rows=job.total_rows,
        processed_rows=job.processed_rows,
        success_count=job.success_count,
        failed_count=job.failed_count,
        percentage_completed=percentage,
        estimated_time_remaining_seconds=job.eta_seconds,
        errors=errors,
        failed_report_download_url=failed_report_download_url,
    )


@router.get("/import-errors/{job_id}/download")
def download_import_errors(
    job_id: str,
    current_user: User = Depends(get_current_admin_or_school_it),
    db: Session = Depends(get_db),
):
    """Download the Excel file that lists rows the worker could not import."""
    repo = ImportRepository(db)
    job = repo.get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Import job not found")
    if job.created_by_user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Import job not found")

    if not job.failed_report_path:
        raise HTTPException(status_code=404, detail="No failed row report available for this job")

    if not os.path.exists(job.failed_report_path):
        raise HTTPException(status_code=404, detail="Failed row report file no longer exists")

    return FileResponse(
        path=job.failed_report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"import_{job.id}_failed_rows.xlsx",
    )

