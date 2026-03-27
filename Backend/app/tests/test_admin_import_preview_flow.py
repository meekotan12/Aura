"""Use: Tests the preview-first bulk student import flow.
Where to use: Run this with pytest to verify preview owns validation and import only queues approved previews.
Role: Test layer. It protects the preview-first import contract from regressions.
"""

from __future__ import annotations

import csv
import io
import json

from openpyxl import Workbook, load_workbook

from app.core.security import create_access_token
from app.models import BulkImportJob, Department, Program, Role, School, StudentProfile, User, UserRole


EXPECTED_HEADERS = [
    "Student_ID",
    "Email",
    "Last Name",
    "First Name",
    "Middle Name",
    "Department",
    "Course",
]


def _create_school(test_db, *, code: str) -> School:
    school = School(
        name=f"Import School {code}",
        school_name=f"Import School {code}",
        school_code=code,
        address="Import Test Address",
    )
    test_db.add(school)
    test_db.commit()
    return school


def _get_or_create_role(test_db, *, name: str) -> Role:
    role = test_db.query(Role).filter(Role.name == name).first()
    if role is None:
        role = Role(name=name)
        test_db.add(role)
        test_db.commit()
    return role


def _create_user_with_role(
    test_db,
    *,
    email: str,
    role_name: str,
    password: str,
    school_id: int | None = None,
    first_name: str = "Import",
    last_name: str = "Tester",
) -> User:
    role = _get_or_create_role(test_db, name=role_name)
    user = User(
        email=email,
        school_id=school_id,
        first_name=first_name,
        last_name=last_name,
        must_change_password=False,
    )
    user.set_password(password)
    test_db.add(user)
    test_db.commit()

    test_db.add(UserRole(user_id=user.id, role_id=role.id))
    test_db.commit()
    return user


def _auth_headers(user: User) -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token({'sub': user.email})}"}


def _create_department_and_program(test_db, *, school_id: int) -> tuple[Department, Program]:
    department = Department(school_id=school_id, name="Computer Science")
    program = Program(school_id=school_id, name="BS Computer Science")
    program.departments.append(department)
    test_db.add_all([department, program])
    test_db.commit()
    return department, program


def _build_workbook_bytes(rows: list[list[str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    sheet.append(EXPECTED_HEADERS)
    for row in rows:
        sheet.append(row)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _build_csv_bytes(rows: list[list[str]]) -> bytes:
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerow(EXPECTED_HEADERS)
    writer.writerows(rows)
    return output.getvalue().encode("utf-8")


def test_preview_import_students_reports_existing_database_conflicts(
    client,
    test_db,
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("IMPORT_STORAGE_DIR", str(tmp_path))

    school = _create_school(test_db, code="IMPORT-PREVIEW-CONFLICT")
    department, program = _create_department_and_program(test_db, school_id=school.id)
    campus_admin = _create_user_with_role(
        test_db,
        email="campus.preview.conflict@example.com",
        role_name="campus_admin",
        password="CampusPass123!",
        school_id=school.id,
    )
    existing_student = _create_user_with_role(
        test_db,
        email="existing.student@example.edu",
        role_name="student",
        password="StudentPass123!",
        school_id=school.id,
        first_name="Existing",
        last_name="Student",
    )
    test_db.add(
        StudentProfile(
            user_id=existing_student.id,
            school_id=school.id,
            student_id="STU-00001",
            department_id=department.id,
            program_id=program.id,
            year_level=1,
        )
    )
    test_db.commit()

    workbook_bytes = _build_workbook_bytes(
        [
            [
                "STU-00001",
                "existing.student@example.edu",
                "Doe",
                "Jane",
                "A",
                "Computer Science",
                "BS Computer Science",
            ]
        ]
    )

    response = client.post(
        "/api/admin/import-students/preview",
        headers=_auth_headers(campus_admin),
        files={
            "file": (
                "students.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["can_commit"] is False
    assert payload["valid_rows"] == 0
    assert payload["invalid_rows"] == 1
    assert payload["preview_token"]
    assert payload["rows"][0]["status"] == "failed"
    assert "Email already exists" in payload["rows"][0]["errors"]
    assert "Duplicate Student_ID within School_ID" in payload["rows"][0]["errors"]

    preview_token = payload["preview_token"]

    preview_error_report = client.get(
        f"/api/admin/import-preview-errors/{preview_token}/download",
        headers=_auth_headers(campus_admin),
    )
    assert preview_error_report.status_code == 200
    report_workbook = load_workbook(io.BytesIO(preview_error_report.content))
    report_sheet = report_workbook.active
    assert report_sheet.max_row == 2
    assert report_sheet.cell(row=2, column=1).value == "STU-00001"
    assert "Email already exists" in str(report_sheet.cell(row=2, column=8).value)
    report_workbook.close()

    preview_retry_file = client.get(
        f"/api/admin/import-preview-errors/{preview_token}/retry-download",
        headers=_auth_headers(campus_admin),
    )
    assert preview_retry_file.status_code == 200
    retry_workbook = load_workbook(io.BytesIO(preview_retry_file.content))
    retry_sheet = retry_workbook.active
    assert retry_sheet.max_row == 2
    assert retry_sheet.cell(row=2, column=1).value == "STU-00001"
    assert retry_sheet.cell(row=2, column=2).value == "existing.student@example.edu"
    retry_workbook.close()

    invalid_import_response = client.post(
        "/api/admin/import-students",
        headers=_auth_headers(campus_admin),
        data={"preview_token": preview_token},
    )
    assert invalid_import_response.status_code == 400
    assert invalid_import_response.json()["detail"] == (
        "Preview still has invalid rows. Fix them before importing."
    )


def test_import_students_requires_preview_token_and_queues_preview_manifest_job(
    client,
    test_db,
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("IMPORT_STORAGE_DIR", str(tmp_path))

    school = _create_school(test_db, code="IMPORT-PREVIEW-QUEUE")
    _create_department_and_program(test_db, school_id=school.id)
    campus_admin = _create_user_with_role(
        test_db,
        email="campus.preview.queue@example.com",
        role_name="campus_admin",
        password="CampusPass123!",
        school_id=school.id,
    )

    workbook_bytes = _build_workbook_bytes(
        [
            [
                "STU-00002",
                "queued.student@example.edu",
                "Queue",
                "Student",
                "B",
                "Computer Science",
                "BS Computer Science",
            ]
        ]
    )

    direct_import_response = client.post(
        "/api/admin/import-students",
        headers=_auth_headers(campus_admin),
        files={
            "file": (
                "students.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert direct_import_response.status_code == 400
    assert direct_import_response.json()["detail"] == "Preview the file first before importing."

    preview_response = client.post(
        "/api/admin/import-students/preview",
        headers=_auth_headers(campus_admin),
        files={
            "file": (
                "students.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["can_commit"] is True
    assert preview_payload["preview_token"]

    preview_token = preview_payload["preview_token"]
    preview_manifest_path = tmp_path / "previews" / f"{preview_token}.json"
    assert preview_manifest_path.exists() is True
    manifest_payload = json.loads(preview_manifest_path.read_text(encoding="utf-8"))
    assert manifest_payload["preview_token"] == preview_token
    assert manifest_payload["target_school_id"] == school.id
    assert len(manifest_payload["rows"]) == 1

    queued_tasks: list[tuple[str, list[str]]] = []

    def _fake_send_task(name: str, args: list[str]):
        queued_tasks.append((name, args))
        return None

    monkeypatch.setattr("app.routers.admin_import.celery_app.send_task", _fake_send_task)

    import_response = client.post(
        "/api/admin/import-students",
        headers=_auth_headers(campus_admin),
        data={"preview_token": preview_token},
    )

    assert import_response.status_code == 200
    import_payload = import_response.json()
    assert import_payload["status"] == "pending"
    assert import_payload["job_id"]

    job = (
        test_db.query(BulkImportJob)
        .filter(BulkImportJob.id == import_payload["job_id"])
        .first()
    )
    assert job is not None
    assert job.original_filename == "students.xlsx"
    assert job.stored_file_path == str(preview_manifest_path)
    assert job.target_school_id == school.id
    assert queued_tasks == [
        ("app.workers.tasks.process_student_import_job", [import_payload["job_id"]])
    ]


def test_preview_import_students_accepts_csv_uploads(
    client,
    test_db,
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("IMPORT_STORAGE_DIR", str(tmp_path))

    school = _create_school(test_db, code="IMPORT-PREVIEW-CSV")
    campus_admin = _create_user_with_role(
        test_db,
        email="campus.preview.csv@example.com",
        role_name="campus_admin",
        password="CampusPass123!",
        school_id=school.id,
    )

    csv_bytes = _build_csv_bytes(
        [
            [
                "STU-00030",
                "csv.student@example.edu",
                "CSV",
                "Student",
                "D",
                "Data Science",
                "BS Data Science",
            ]
        ]
    )

    response = client.post(
        "/api/admin/import-students/preview",
        headers=_auth_headers(campus_admin),
        files={"file": ("students.csv", csv_bytes, "text/csv")},
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["can_commit"] is True
    assert payload["valid_rows"] == 1
    assert payload["invalid_rows"] == 0
    assert payload["preview_token"]


def test_remove_invalid_preview_rows_keeps_only_valid_rows_and_allows_import(
    client,
    test_db,
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("IMPORT_STORAGE_DIR", str(tmp_path))

    school = _create_school(test_db, code="IMPORT-PREVIEW-CLEAN")
    department, program = _create_department_and_program(test_db, school_id=school.id)
    campus_admin = _create_user_with_role(
        test_db,
        email="campus.preview.clean@example.com",
        role_name="campus_admin",
        password="CampusPass123!",
        school_id=school.id,
    )
    existing_student = _create_user_with_role(
        test_db,
        email="existing.clean@example.edu",
        role_name="student",
        password="StudentPass123!",
        school_id=school.id,
        first_name="Existing",
        last_name="Student",
    )
    test_db.add(
        StudentProfile(
            user_id=existing_student.id,
            school_id=school.id,
            student_id="STU-00010",
            department_id=department.id,
            program_id=program.id,
            year_level=1,
        )
    )
    test_db.commit()

    workbook_bytes = _build_workbook_bytes(
        [
            [
                "STU-00010",
                "existing.clean@example.edu",
                "Duplicate",
                "Student",
                "A",
                "Computer Science",
                "BS Computer Science",
            ],
            [
                "STU-00011",
                "valid.clean@example.edu",
                "Valid",
                "Student",
                "B",
                "Computer Science",
                "BS Computer Science",
            ],
        ]
    )

    preview_response = client.post(
        "/api/admin/import-students/preview",
        headers=_auth_headers(campus_admin),
        files={
            "file": (
                "students.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert preview_response.status_code == 200
    preview_payload = preview_response.json()
    assert preview_payload["can_commit"] is False
    assert preview_payload["valid_rows"] == 1
    assert preview_payload["invalid_rows"] == 1
    preview_token = preview_payload["preview_token"]
    assert preview_token

    cleaned_response = client.post(
        f"/api/admin/import-preview-errors/{preview_token}/remove-invalid",
        headers=_auth_headers(campus_admin),
    )

    assert cleaned_response.status_code == 200
    cleaned_payload = cleaned_response.json()
    assert cleaned_payload["preview_token"] == preview_token
    assert cleaned_payload["total_rows"] == 1
    assert cleaned_payload["valid_rows"] == 1
    assert cleaned_payload["invalid_rows"] == 0
    assert cleaned_payload["can_commit"] is True
    assert len(cleaned_payload["rows"]) == 1
    assert cleaned_payload["rows"][0]["status"] == "valid"
    assert cleaned_payload["rows"][0]["row"] == 3
    assert cleaned_payload["rows"][0]["row_data"]["Email"] == "valid.clean@example.edu"

    preview_manifest_path = tmp_path / "previews" / f"{preview_token}.json"
    manifest_payload = json.loads(preview_manifest_path.read_text(encoding="utf-8"))
    assert manifest_payload["total_rows"] == 1
    assert manifest_payload["valid_rows"] == 1
    assert manifest_payload["invalid_rows"] == 0
    assert manifest_payload["can_commit"] is True
    assert manifest_payload["error_rows"] == []
    assert len(manifest_payload["rows"]) == 1

    queued_tasks: list[tuple[str, list[str]]] = []

    def _fake_send_task(name: str, args: list[str]):
        queued_tasks.append((name, args))
        return None

    monkeypatch.setattr("app.routers.admin_import.celery_app.send_task", _fake_send_task)

    import_response = client.post(
        "/api/admin/import-students",
        headers=_auth_headers(campus_admin),
        data={"preview_token": preview_token},
    )

    assert import_response.status_code == 200
    import_payload = import_response.json()
    assert import_payload["status"] == "pending"
    assert queued_tasks == [
        ("app.workers.tasks.process_student_import_job", [import_payload["job_id"]])
    ]


def test_import_students_falls_back_to_in_process_job_when_celery_dispatch_fails(
    client,
    test_db,
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("IMPORT_STORAGE_DIR", str(tmp_path))

    school = _create_school(test_db, code="IMPORT-PREVIEW-FALLBACK")
    _create_department_and_program(test_db, school_id=school.id)
    campus_admin = _create_user_with_role(
        test_db,
        email="campus.preview.fallback@example.com",
        role_name="campus_admin",
        password="CampusPass123!",
        school_id=school.id,
    )

    workbook_bytes = _build_workbook_bytes(
        [
            [
                "STU-00020",
                "fallback.student@example.edu",
                "Fallback",
                "Student",
                "B",
                "Computer Science",
                "BS Computer Science",
            ]
        ]
    )

    preview_response = client.post(
        "/api/admin/import-students/preview",
        headers=_auth_headers(campus_admin),
        files={
            "file": (
                "students.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert preview_response.status_code == 200
    preview_token = preview_response.json()["preview_token"]
    assert preview_token

    fallback_runs: list[str] = []

    def _raise_send_task(_name: str, *args, **kwargs):
        raise RuntimeError("celery broker unavailable")

    def _fake_process_job(self, job_id: str):
        fallback_runs.append(job_id)

    monkeypatch.setattr("app.routers.admin_import.celery_app.send_task", _raise_send_task)
    monkeypatch.setattr("app.services.student_import_service.celery_app.send_task", _raise_send_task)
    monkeypatch.setattr("app.routers.admin_import.StudentImportService.process_job", _fake_process_job)

    import_response = client.post(
        "/api/admin/import-students",
        headers=_auth_headers(campus_admin),
        data={"preview_token": preview_token},
    )

    assert import_response.status_code == 200
    import_payload = import_response.json()
    assert import_payload["status"] == "pending"

    job = (
        test_db.query(BulkImportJob)
        .filter(BulkImportJob.id == import_payload["job_id"])
        .first()
    )
    assert job is not None
    assert job.status == "pending"
    assert fallback_runs == [import_payload["job_id"]]


def test_preview_import_students_accepts_new_department_program_pairings(
    client,
    test_db,
    tmp_path,
    monkeypatch,
):
    monkeypatch.setenv("IMPORT_STORAGE_DIR", str(tmp_path))

    school = _create_school(test_db, code="IMPORT-PREVIEW-MISMATCH")
    _create_department_and_program(test_db, school_id=school.id)
    mismatched_department = Department(school_id=school.id, name="Information Technology")
    test_db.add(mismatched_department)
    test_db.commit()

    campus_admin = _create_user_with_role(
        test_db,
        email="campus.preview.mismatch@example.com",
        role_name="campus_admin",
        password="CampusPass123!",
        school_id=school.id,
    )

    workbook_bytes = _build_workbook_bytes(
        [
            [
                "STU-00003",
                "mismatch.student@example.edu",
                "Mismatch",
                "Student",
                "C",
                "Information Technology",
                "BS Computer Science",
            ]
        ]
    )

    response = client.post(
        "/api/admin/import-students/preview",
        headers=_auth_headers(campus_admin),
        files={
            "file": (
                "students.xlsx",
                workbook_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["can_commit"] is True
    assert payload["valid_rows"] == 1
    assert payload["invalid_rows"] == 0
    assert payload["rows"][0]["status"] == "valid"
