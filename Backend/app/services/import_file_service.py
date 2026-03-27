"""Use: Normalizes uploaded import files into CSV-backed row data.
Where to use: Use this from import routes or workers when handling `.csv` and `.xlsx` student import files.
Role: Service layer. It keeps file-format handling separate from validation and database work.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from openpyxl import load_workbook

SUPPORTED_IMPORT_FILE_EXTENSIONS = {".csv", ".xlsx"}


def get_import_file_extension(filename: str) -> str:
    return Path(filename or "").suffix.strip().lower()


def is_supported_import_file(filename: str) -> bool:
    return get_import_file_extension(filename) in SUPPORTED_IMPORT_FILE_EXTENSIONS


def _xlsx_bytes_to_csv_text(file_bytes: bytes) -> str:
    workbook = load_workbook(filename=io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow("" if value is None else value for value in row)
        return output.getvalue()
    finally:
        workbook.close()


def _csv_text_from_bytes(file_bytes: bytes) -> str:
    return file_bytes.decode("utf-8-sig")


def load_tabular_rows_from_bytes(*, filename: str, file_bytes: bytes) -> list[list[str]]:
    extension = get_import_file_extension(filename)
    if extension == ".csv":
        csv_text = _csv_text_from_bytes(file_bytes)
    elif extension == ".xlsx":
        csv_text = _xlsx_bytes_to_csv_text(file_bytes)
    else:
        raise ValueError("Unsupported import file type")

    return [list(row) for row in csv.reader(io.StringIO(csv_text))]


def normalize_upload_to_csv_bytes(*, filename: str, file_bytes: bytes) -> tuple[str, bytes]:
    rows = load_tabular_rows_from_bytes(filename=filename, file_bytes=file_bytes)
    output = io.StringIO(newline="")
    writer = csv.writer(output)
    writer.writerows(rows)
    normalized_filename = f"{Path(filename or 'student_import').stem}.csv"
    return normalized_filename, output.getvalue().encode("utf-8")
